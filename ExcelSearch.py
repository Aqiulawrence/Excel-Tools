"""Excel workbook text search tool."""

import concurrent.futures
import json
import multiprocessing
import os
import sys
import warnings
import winreg
from dataclasses import dataclass
from itertools import chain, islice
from typing import Mapping, Sequence

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from PyQt6.QtCore import QSettings, QThread, QTimer, Qt, pyqtSignal
from PyQt6.QtGui import QColor
from PyQt6.QtWidgets import (
    QApplication, QDialog, QDialogButtonBox, QFileDialog, QFormLayout,
    QGroupBox, QHBoxLayout, QHeaderView, QLabel, QLineEdit, QMainWindow,
    QMenu, QMessageBox, QProgressBar, QPushButton, QSpinBox,
    QStyleFactory, QTableWidget, QTableWidgetItem, QVBoxLayout, QWidget,
)
from xlrd import open_workbook

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

VERSION = "2.0"
APP_NAME = "Excel价格搜索工具 by Sam"
COMPANY_NAME = "Sam"
HEADER_SCAN_ROWS = 30
DEFAULT_MAX_WORKERS = min(8, os.cpu_count() or 4)
EXCEL_EXTENSIONS = (".xls", ".xlsx", ".xlsm")
DEFAULT_KEYWORDS = {
    "description": ["品名", "description", "desc"],
    "rmb": ["RMB", "人民币", "￥", "¥"],
    "usd": ["USD", "$", "美元"],
}
RESULT_FIELDS = (
    ("description", "description"),
    ("rmb", "rmb_price"),
    ("usd", "usd_price"),
)


@dataclass
class SearchResult:
    file_path: str
    sheet_name: str
    cell_address: str
    cell_value: str
    description: str = ""
    rmb_price: str = ""
    usd_price: str = ""


def clean_keywords(keywords: Mapping[str, Sequence[str]]) -> dict[str, list[str]]:
    """Normalize settings and ensure every supported field is present."""
    cleaned = {}
    for name in DEFAULT_KEYWORDS:
        values = []
        for word in keywords.get(name, []):
            value = str(word).strip()
            if value:
                values.append(value)
        cleaned[name] = values
    return cleaned


def find_keyword_columns(rows, keywords):
    """Locate the first matching header column for each output field."""
    columns = dict.fromkeys(DEFAULT_KEYWORDS)
    normalized = {
        name: [word.lower() for word in words]
        for name, words in keywords.items()
    }

    for row in rows:
        for index, value in enumerate(row):
            if value is None:
                continue

            text = str(value).lower()
            for name, words in normalized.items():
                if columns[name] is None and any(word in text for word in words):
                    columns[name] = index
    return columns


def build_result(file_path, sheet_name, row_number, column, value, row, columns):
    """Create one result and copy values from the detected columns."""
    details = {}
    for keyword, field in RESULT_FIELDS:
        index = columns[keyword]
        if index is not None and index < len(row) and row[index] is not None:
            details[field] = str(row[index])
    address = f"{get_column_letter(column + 1)}{row_number}"
    return SearchResult(file_path, sheet_name, address, str(value), **details)


def search_rows(file_path, sheet_name, rows, search_term, keywords, cancel_event):
    """Search all rows in one worksheet and return matching cells."""
    iterator = iter(rows)

    # Headers are normally near the top. A bounded scan avoids treating a
    # matching value deep in the sheet as a column header.
    headers = list(islice(iterator, HEADER_SCAN_ROWS))
    columns = find_keyword_columns(headers, keywords)
    matches = []

    for row_number, row in enumerate(chain(headers, iterator), 1):
        if cancel_event.is_set():
            return None

        for column, value in enumerate(row):
            if value is not None and search_term in str(value).lower():
                matches.append(
                    build_result(
                        file_path,
                        sheet_name,
                        row_number,
                        column,
                        value,
                        row,
                        columns,
                    )
                )

    return matches


def search_file(file_path, search_term, keywords, cancel_event):
    """Search one workbook in a separate process; False means it could not be read."""
    try:
        if file_path.lower().endswith(".xls"):
            # xlrd is used for the legacy binary format.
            workbook = open_workbook(file_path)

            worksheets = (
                (sheet.name, (sheet.row_values(i) for i in range(sheet.nrows)))
                for sheet in workbook.sheets()
            )
            close = None
        else:
            # openpyxl handles modern workbooks in read-only mode.
            workbook = load_workbook(file_path, read_only=True, data_only=True)

            worksheets = (
                (sheet.title, sheet.iter_rows(values_only=True))
                for sheet in workbook.worksheets
            )
            close = workbook.close

        try:
            results = []
            for sheet_name, rows in worksheets:
                matches = search_rows(
                    file_path,
                    sheet_name,
                    rows,
                    search_term,
                    keywords,
                    cancel_event,
                )
                if matches is None:
                    return [], True
                results.extend(matches)

            return results, True
        finally:
            if close is not None:
                close()
    except Exception:
        return [], False


class KeywordSettingsDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("关键词设置")

        layout = QVBoxLayout(self)
        form = QFormLayout()
        self.inputs = {name: QLineEdit() for name in DEFAULT_KEYWORDS}

        fields = (
            ("description", "品名关键词"),
            ("rmb", "RMB关键词"),
            ("usd", "USD关键词"),
        )
        for name, label in fields:
            form.addRow(label, self.inputs[name])
        layout.addLayout(form)

        buttons = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok
            | QDialogButtonBox.StandardButton.Cancel
            | QDialogButtonBox.StandardButton.Reset
        )
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        buttons.button(QDialogButtonBox.StandardButton.Reset).clicked.connect(self.reset_default)
        layout.addWidget(buttons)

    def reset_default(self):
        self.set_keywords(DEFAULT_KEYWORDS)

    def get_keywords(self):
        values = {
            name: input_.text().split(",")
            for name, input_ in self.inputs.items()
        }
        return clean_keywords(values)

    def set_keywords(self, keywords):
        for name, input_ in self.inputs.items():
            input_.setText(",".join(keywords.get(name, [])))


class ExcelSearchWorker(QThread):
    progress_signal = pyqtSignal(int, int, int)
    batch_result_signal = pyqtSignal(list)
    finished_signal = pyqtSignal(int, int, int)

    def __init__(self):
        super().__init__()
        self.search_term = ""
        self.search_paths: list[str] = []
        self.max_workers = DEFAULT_MAX_WORKERS
        self.keywords = clean_keywords(DEFAULT_KEYWORDS)
        self._running = False
        self._cancel_event = None

    def setup(self, search_term, search_paths, max_workers, keywords):
        self.search_term = search_term.lower().strip()
        self.search_paths = list(search_paths)
        self.max_workers = max_workers
        self.keywords = clean_keywords(keywords)

    def run(self):
        files = list(self.collect_files())
        total_files = len(files)
        completed = succeeded = failed = found = 0
        self._running = True

        with multiprocessing.Manager() as manager:
            self._cancel_event = manager.Event()
            with concurrent.futures.ProcessPoolExecutor(max_workers=self.max_workers) as executor:
                pending = iter(files)
                futures = set()

                def submit_next():
                    """Keep at most max_workers files in flight."""
                    try:
                        file_path = next(pending)
                    except StopIteration:
                        return False

                    future = executor.submit(
                        search_file,
                        file_path,
                        self.search_term,
                        self.keywords,
                        self._cancel_event,
                    )
                    futures.add(future)
                    return True

                # Submit an initial batch, then replenish it as files finish.
                while self._running and len(futures) < self.max_workers:
                    if not submit_next():
                        break

                while futures and self._running:
                    done, _ = concurrent.futures.wait(
                        futures,
                        return_when=concurrent.futures.FIRST_COMPLETED,
                    )
                    for future in done:
                        futures.remove(future)
                        completed += 1

                        try:
                            results, success = future.result()
                            if success:
                                succeeded += 1
                                found += len(results)
                                if results:
                                    self.batch_result_signal.emit(results)
                            else:
                                failed += 1
                        except Exception:
                            failed += 1

                        self.progress_signal.emit(completed, total_files, found)

                    while self._running and len(futures) < self.max_workers:
                        if not submit_next():
                            break

                for future in futures:
                    future.cancel()

        self._cancel_event = None
        self._running = False
        self.finished_signal.emit(succeeded, failed, found)

    def collect_files(self):
        for path in self.search_paths:
            if os.path.isdir(path):
                for root, _, filenames in os.walk(path):
                    for filename in filenames:
                        is_excel = filename.lower().endswith(EXCEL_EXTENSIONS)
                        is_temporary = filename.startswith(("~$", "$"))
                        if is_excel and not is_temporary:
                            yield os.path.join(root, filename)

    def stop(self):
        self._running = False
        if self._cancel_event is not None:
            self._cancel_event.set()


class ResultTableWidget(QTableWidget):
    HEADERS = ("文件", "单元格", "单元格内容", "品名", "RMB 价格", "USD 价格")
    PRICE_COLORS = {4: "#0070c0", 5: "#c00000"}

    def __init__(self):
        super().__init__()
        self.results: list[SearchResult] = []

        self.setColumnCount(len(self.HEADERS))
        self.setHorizontalHeaderLabels(self.HEADERS)
        self.configure_columns()
        self.setAlternatingRowColors(True)
        self.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)

        self.doubleClicked.connect(self.open_excel_file)
        self.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.customContextMenuRequested.connect(self.show_context_menu)

    def configure_columns(self):
        header = self.horizontalHeader()

        for column in (0, 2, 3):
            header.setSectionResizeMode(column, QHeaderView.ResizeMode.Stretch)

        for column, width in ((1, 60), (4, 80), (5, 80)):
            header.setSectionResizeMode(column, QHeaderView.ResizeMode.Fixed)
            self.setColumnWidth(column, width)

    @staticmethod
    def text_item(value):
        item = QTableWidgetItem(value)
        item.setToolTip(value)
        return item

    def add_batch_results(self, results):
        start_row = self.rowCount()
        self.results.extend(results)
        self.setRowCount(start_row + len(results))

        for offset, result in enumerate(results):
            row = start_row + offset
            file_item = self.text_item(os.path.basename(result.file_path))
            file_item.setToolTip(result.file_path)
            self.setItem(row, 0, file_item)
            values = (result.cell_address, result.cell_value, result.description)
            for column, value in enumerate(values, 1):
                self.setItem(row, column, self.text_item(value))

            self.set_price_item(row, 4, result.rmb_price, self.PRICE_COLORS[4])
            self.set_price_item(row, 5, result.usd_price, self.PRICE_COLORS[5])

    def set_price_item(self, row, column, price, color):
        item = self.text_item(price)
        if price and any(character.isdigit() for character in price):
            item.setForeground(QColor(color))
        self.setItem(row, column, item)

    def clear_results(self):
        self.results.clear()
        self.setRowCount(0)

    def selected_result(self):
        row = self.currentRow()
        return self.results[row] if 0 <= row < len(self.results) else None

    def open_excel_file(self):
        if result := self.selected_result():
            if os.path.exists(result.file_path):
                os.startfile(result.file_path)

    def show_context_menu(self, position):
        menu = QMenu(self)
        copy_action = menu.addAction("复制")
        copy_action.triggered.connect(self.copy_cell_content)
        menu.addSeparator()
        open_file_action = menu.addAction("打开文件")
        open_file_action.triggered.connect(self.open_excel_file)
        location_action = menu.addAction("打开文件所在目录")
        location_action.triggered.connect(self.open_file_location)
        menu.exec(self.mapToGlobal(position))

    def copy_cell_content(self):
        if item := self.currentItem():
            QApplication.clipboard().setText(item.text())

    def open_file_location(self):
        for row in {item.row() for item in self.selectedItems()}:
            if row < len(self.results):
                os.startfile(os.path.dirname(self.results[row].file_path))


class ExcelSearchTool(QMainWindow):
    def __init__(self):
        super().__init__()
        self.search_worker: ExcelSearchWorker | None = None
        self.settings = QSettings(COMPANY_NAME, APP_NAME)
        self.search_history: list[str] = []
        self.custom_keywords = clean_keywords(DEFAULT_KEYWORDS)
        self.setWindowTitle(f"{APP_NAME} v{VERSION}")
        self.resize(700, 550)
        self.init_ui()
        self.load_settings()
        self.center_window()

    def init_ui(self):
        central = QWidget()
        self.setCentralWidget(central)
        layout = QVBoxLayout(central)

        layout.addWidget(self.create_search_group())
        self.progress_bar = QProgressBar()
        self.progress_bar.setFormat("就绪")
        layout.addWidget(self.progress_bar)

        group = QGroupBox("搜索结果")
        result_layout = QVBoxLayout(group)
        self.result_table = ResultTableWidget()
        result_layout.addWidget(self.result_table)
        layout.addWidget(group)

    def create_search_group(self):
        group = QGroupBox("搜索设置")
        layout = QVBoxLayout(group)

        layout.addLayout(self.create_keyword_row())
        layout.addLayout(self.create_folder_row())
        layout.addLayout(self.create_action_row())
        return group

    def create_keyword_row(self):
        row = QHBoxLayout()
        row.addWidget(QLabel("搜索关键词"))

        self.keyword_input = QLineEdit()
        self.keyword_input.returnPressed.connect(self.start_search)
        row.addWidget(self.keyword_input)

        for label, callback in (
            ("历史", self.show_history),
            ("关键词设置", self.show_keyword_settings),
        ):
            button = QPushButton(label)
            button.clicked.connect(callback)
            row.addWidget(button)

        return row

    def create_folder_row(self):
        row = QHBoxLayout()
        row.addWidget(QLabel("搜索文件夹"))

        self.folder_input = QLineEdit()
        self.folder_input.setReadOnly(True)
        row.addWidget(self.folder_input)

        folder_button = QPushButton("选择文件夹")
        folder_button.clicked.connect(self.select_folder)
        row.addWidget(folder_button)

        return row

    def create_action_row(self):
        row = QHBoxLayout()
        row.addWidget(QLabel("并发文件数"))

        self.thread_spin = QSpinBox()
        self.thread_spin.setRange(1, 32)
        self.thread_spin.setValue(DEFAULT_MAX_WORKERS)
        row.addWidget(self.thread_spin)
        row.addStretch()

        self.search_btn = QPushButton("开始搜索")
        self.search_btn.clicked.connect(self.start_search)
        row.addWidget(self.search_btn)

        self.stop_btn = QPushButton("停止搜索")
        self.stop_btn.clicked.connect(self.stop_search)
        self.stop_btn.setEnabled(False)
        row.addWidget(self.stop_btn)

        return row

    def center_window(self):
        screen = QApplication.primaryScreen().availableGeometry()
        self.move(screen.center() - self.rect().center())

    def load_settings(self):
        if path := self.settings.value("last_search_path", ""):
            if os.path.isdir(path):
                self.folder_input.setText(path)
        thread_count = self.settings.value(
            "thread_count",
            DEFAULT_MAX_WORKERS,
            type=int,
        )
        self.thread_spin.setValue(thread_count)
        self.search_history = list(self.settings.value("search_history", []))

        try:
            saved_keywords = json.loads(
                self.settings.value("custom_keywords", "")
            )
            self.custom_keywords = clean_keywords(saved_keywords)
        except (TypeError, json.JSONDecodeError):
            pass

    def save_settings(self):
        self.settings.setValue("thread_count", self.thread_spin.value())
        self.settings.setValue("search_history", self.search_history[-20:])
        self.settings.setValue(
            "custom_keywords",
            json.dumps(self.custom_keywords, ensure_ascii=False),
        )

    def select_folder(self):
        if folder := QFileDialog.getExistingDirectory(self, "选择文件夹"):
            self.folder_input.setText(folder)
            self.settings.setValue("last_search_path", folder)

    def show_keyword_settings(self):
        dialog = KeywordSettingsDialog(self)
        dialog.set_keywords(self.custom_keywords)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            self.custom_keywords = dialog.get_keywords()
            self.save_settings()

    def start_search(self):
        keyword = self.keyword_input.text().strip()
        folder = self.folder_input.text().strip()
        if not keyword:
            QMessageBox.warning(self, "提示", "请输入搜索关键词")
            return
        if not os.path.isdir(folder):
            QMessageBox.warning(self, "提示", "请选择有效的文件夹")
            return
        if keyword not in self.search_history:
            self.search_history.append(keyword)
        self.result_table.clear_results()
        self.search_btn.setEnabled(False)
        self.stop_btn.setEnabled(True)
        self.progress_bar.setValue(0)
        self.progress_bar.setFormat("准备搜索…")
        self.search_worker = ExcelSearchWorker()
        self.search_worker.setup(
            keyword,
            [folder],
            self.thread_spin.value(),
            self.custom_keywords,
        )
        self.search_worker.progress_signal.connect(self.update_progress)
        self.search_worker.batch_result_signal.connect(self.result_table.add_batch_results)
        self.search_worker.finished_signal.connect(self.search_finished)
        self.search_worker.start()

    def stop_search(self):
        if self.search_worker and self.search_worker.isRunning():
            self.search_worker.stop()
            self.stop_btn.setEnabled(False)

    def update_progress(self, current, total, found):
        percentage = int(current / total * 100) if total else 0
        self.progress_bar.setValue(percentage)
        self.progress_bar.setFormat(
            f"正在搜索：{current}/{total} ({percentage}%) — 已找到 {found} 条"
        )

    def search_finished(self, success, failed, found):
        self.search_btn.setEnabled(True)
        self.stop_btn.setEnabled(False)
        self.progress_bar.setValue(100)
        self.progress_bar.setFormat(f"搜索完成：找到 {found} 条（成功：{success}，失败：{failed}）")
        if found:
            QApplication.beep()
        self.save_settings()

    def show_history(self):
        if not self.search_history:
            return
        menu = QMenu(self)
        for keyword in reversed(self.search_history[-10:]):
            action = menu.addAction(keyword)
            action.triggered.connect(lambda _, value=keyword: self.set_and_search(value))
        menu.addSeparator()
        clear_action = menu.addAction("清空历史")
        clear_action.triggered.connect(self.clear_history)
        menu.exec(self.cursor().pos())

    def set_and_search(self, keyword):
        self.keyword_input.setText(keyword)
        QTimer.singleShot(100, self.start_search)

    def clear_history(self):
        self.search_history.clear()
        self.save_settings()

    def closeEvent(self, event):
        if self.search_worker and self.search_worker.isRunning():
            self.search_worker.stop()
            self.search_worker.wait(2_000)
        self.save_settings()
        event.accept()


def register_updater_startup():
    """Keep the original updater auto-start behavior when the updater is present."""
    updater_path = os.path.join(os.path.dirname(sys.argv[0]), "Updater.exe")
    if not os.path.isfile(updater_path):
        return

    run_key = r"Software\Microsoft\Windows\CurrentVersion\Run"
    with winreg.OpenKey(
        winreg.HKEY_CURRENT_USER,
        run_key,
        0,
        winreg.KEY_SET_VALUE,
    ) as key:
        winreg.SetValueEx(key, "Updater", 0, winreg.REG_SZ, f'"{updater_path}"')


def main():
    multiprocessing.freeze_support()
    app = QApplication(sys.argv)
    app.setApplicationName(APP_NAME)
    app.setOrganizationName(COMPANY_NAME)
    app.setStyle(QStyleFactory.create("Fusion"))
    window = ExcelSearchTool()
    window.show()
    register_updater_startup()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
