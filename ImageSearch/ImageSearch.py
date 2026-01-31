# main.py
import sys
import os
import shutil
import time
import requests
from urllib.parse import urlparse
from concurrent.futures import ThreadPoolExecutor, as_completed

import openpyxl
from openpyxl.drawing.image import Image
from bs4 import BeautifulSoup
from PyQt6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout,
                             QHBoxLayout, QGridLayout, QGroupBox, QLabel,
                             QLineEdit, QPushButton, QTextEdit, QFileDialog,
                             QMessageBox, QProgressBar, QCheckBox, QSpinBox)
from PyQt6.QtCore import Qt, pyqtSignal, QThread, QTimer
from PyQt6.QtGui import QFont, QTextCursor

# 配置常量
VERSION = "1.0"
IMG_DIR = './imgs'

def extract_excel_data(start_cell, end_cell, excel_file):
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active

    # 解析单元格坐标
    start_col = ''.join(filter(str.isalpha, start_cell))
    start_row = int(''.join(filter(str.isdigit, start_cell)))
    end_col = ''.join(filter(str.isalpha, end_cell))
    end_row = int(''.join(filter(str.isdigit, end_cell)))

    data = []
    for row in range(start_row, end_row + 1):
        row_data = []
        for col in range(ord(start_col) - 64, ord(end_col) - 64 + 1):
            cell = ws.cell(row=row, column=col)
            value = str(cell.value).split('\n')[0] if cell.value else ''
            row_data.append(value)

        if len(row_data) == 1:
            data.append(row_data[0])
        else:
            data.append(row_data)

    return data


def insert_images_to_excel(start_cell, image_count, excel_file):
    if not os.path.exists(IMG_DIR):
        return 0

    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active

    # 解析起始单元格
    col = ''.join(filter(str.isalpha, start_cell))
    row = int(''.join(filter(str.isdigit, start_cell)))

    error_count = 0
    img_names = [f'{i:03d}.png' for i in range(image_count)]

    for index, name in enumerate(img_names):
        img_path = os.path.join(IMG_DIR, name)

        if not os.path.exists(img_path) or os.path.getsize(img_path) == 0:
            error_count += 1
            row += 1
            continue

        try:
            # 获取单元格尺寸
            cell_ref = f"{col}{row}"
            w1 = ws.column_dimensions[col].width
            h1 = ws.row_dimensions[row].height
            w2 = ws.sheet_format.defaultColWidth
            h2 = ws.sheet_format.defaultRowHeight

            # 计算宽度
            if w1 == 13:
                width = w2 if w2 is not None else w1
            else:
                width = w1

            # 计算高度
            height = h2 if h1 is None else h1

            # 转换为像素
            width *= 8
            height *= 1.3

            # 加载并调整图片
            img = Image(img_path)
            original_width, original_height = img.width, img.height

            # 计算缩放比例
            scale_width = width / original_width
            scale_height = height / original_height
            scale = min(scale_width, scale_height)

            # 调整大小
            img.width = int(original_width * scale)
            img.height = int(original_height * scale)

            # 插入图片
            ws.add_image(img, cell_ref)

        except Exception as e:
            error_count += 1
            print(f"插入失败 {index}: {e}")

        row += 1

    # 保存文件
    try:
        wb.save(excel_file)
    except PermissionError as e:
        raise Exception("保存失败！请确保文件未被占用且不是只读文件。")

    return error_count


class ImageSearchWorker(QThread):
    search_started = pyqtSignal(int)  # 总任务数
    item_completed = pyqtSignal(int, bool, str)  # 索引，是否成功，消息
    search_finished = pyqtSignal(list)  # 失败件号列表
    search_error = pyqtSignal(str)  # 错误信息

    def __init__(self):
        super().__init__()
        self.search_terms = []
        self.max_workers = 5
        self.enable_filter = True
        self.priority_sites = ['ebay.com', 'amazon.com', 'cat.com', 'alibaba.com']
        self.blacklist_sites = ['farfetch.com']
        self.headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:50.0) Gecko/20100101 Firefox/50.0'}

    def setup(self, search_terms, max_workers=5, enable_filter=True):
        self.search_terms = [x.strip() for x in search_terms if x.strip()]
        self.max_workers = max_workers
        self.enable_filter = enable_filter

        # 清空并创建图片目录
        if os.path.exists(IMG_DIR):
            shutil.rmtree(IMG_DIR)
        os.makedirs(IMG_DIR)

    def search_website(self, tag):
        span_count = 0
        current_tag = tag

        while current_tag and span_count < 4:
            current_tag = current_tag.find_next()
            if current_tag and current_tag.name == 'span':
                span_count += 1

        if current_tag and span_count == 4:
            return current_tag.get_text(strip=True)
        return ""

    def download_image(self, url, index, term):
        file_name = os.path.join(IMG_DIR, f'{index:03d}.png')

        # 获取网页内容
        try:
            response = requests.get(url, headers=self.headers, timeout=30)
        except Exception as e:
            return False, term, "网络请求失败"

        # 检查是否被Google屏蔽
        if "Our systems have detected unusual traffic" in response.text:
            return False, term, "被Google屏蔽"

        # 解析图片
        soup = BeautifulSoup(response.text, "html.parser")
        img_tags = soup.find_all("img")

        if len(img_tags) > 0:
            del img_tags[0]  # 删除Google logo

        # 筛选图片
        record = []
        if self.enable_filter:
            for img in img_tags:
                website = self.search_website(img)
                for pri in self.priority_sites:
                    if pri in website:
                        record.append(img)

        if not record:
            record = img_tags
            for img in img_tags:
                website = self.search_website(img)
                for black in self.blacklist_sites:
                    if black in website:
                        if img in record:
                            record.remove(img)

        # 下载图片
        for img in record:
            try:
                src = img.attrs.get("src", "")
                if src.startswith("http"):  # 防止base64图片
                    img_response = requests.get(src, headers=self.headers, timeout=30)

                    with open(file_name, 'wb') as f:
                        f.write(img_response.content)

                    return True, term, "成功"

            except Exception:
                continue

        # 创建空文件
        open(file_name, 'w').close()
        return False, term, "未找到图片"

    def run(self):
        try:
            # 测试网络连接
            test_url = "https://www.google.com"
            response = requests.get(test_url, timeout=10)

            if "Our systems have detected unusual traffic" in response.text:
                self.search_error.emit("Google检测到异常流量，请稍后再试")
                return

        except Exception as e:
            self.search_error.emit(f"网络连接失败: {str(e)}")
            return

        # 通知开始搜索
        total_tasks = len(self.search_terms)
        self.search_started.emit(total_tasks)

        failed_items = []

        # 使用线程池并发搜索
        with ThreadPoolExecutor(max_workers=self.max_workers) as executor:
            futures = {}

            # 提交所有任务
            for index, term in enumerate(self.search_terms):
                url = f'https://www.google.com.hk/search?q={term}&udm=2'
                future = executor.submit(self.download_image, url, index, term)
                futures[future] = (index, term)

            # 处理完成的任务
            for future in as_completed(futures):
                index, term = futures[future]
                try:
                    success, item_term, message = future.result(timeout=60)
                    if not success:
                        failed_items.append((index, term, message))
                    self.item_completed.emit(index, success, message)
                except Exception as e:
                    failed_items.append((index, term, f"任务异常: {str(e)}"))
                    self.item_completed.emit(index, False, f"异常: {str(e)}")

        # 发射完成信号
        self.search_finished.emit(failed_items)


# 支持拖拽的文本输入框
class DragDropLineEdit(QLineEdit):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setAcceptDrops(True)

    def dragEnterEvent(self, event):
        if event.mimeData().hasUrls():
            event.accept()
        else:
            event.ignore()

    def dropEvent(self, event):
        files = [u.toLocalFile() for u in event.mimeData().urls()]
        if files:
            # 只取第一个文件
            file_path = files[0]
            if file_path.lower().endswith('.xlsx'):
                self.setText(file_path)
            else:
                QMessageBox.warning(self, "警告", "请拖拽.xlsx格式的Excel文件")


# 主界面
class ExcelToolsGUI(QMainWindow):
    def __init__(self):
        super().__init__()
        self.search_worker = ImageSearchWorker()
        self.completed_count = 0
        self.total_tasks = 0
        self.init_ui()
        self.setup_connections()

    def init_ui(self):
        """初始化界面"""
        self.setWindowTitle(f"件号搜图工具 by Sam v{VERSION}")
        self.resize(0, 700)

        # 设置字体
        font = QFont("Microsoft YaHei", 10)
        self.setFont(font)

        # 中央部件
        central_widget = QWidget()
        self.setCentralWidget(central_widget)

        # 主布局
        main_layout = QVBoxLayout(central_widget)
        main_layout.setSpacing(10)
        main_layout.setContentsMargins(10, 10, 10, 10)

        # 文件选择区域
        file_group = self.create_file_group()
        main_layout.addWidget(file_group)

        # 参数设置区域（统一的参数框）
        param_group = self.create_param_group()
        main_layout.addWidget(param_group)

        # 内容和日志区域（并排放置）
        content_log_group = self.create_content_log_group()
        main_layout.addWidget(content_log_group, 1)  # 设置伸缩因子为1

        # 操作按钮区域
        button_group = self.create_button_group()
        main_layout.addWidget(button_group)

        # 进度条区域（始终显示）
        progress_group = self.create_progress_group()
        main_layout.addWidget(progress_group)

        # 删除状态栏
        self.setStatusBar(None)

    def create_file_group(self):
        """创建文件选择区域"""
        group = QGroupBox("文件设置")
        layout = QHBoxLayout()
        layout.setSpacing(10)

        layout.addWidget(QLabel("Excel文件:"))

        # 使用支持拖拽的自定义控件
        self.file_path_edit = DragDropLineEdit()
        self.file_path_edit.setReadOnly(True)
        layout.addWidget(self.file_path_edit, 1)

        self.select_file_btn = QPushButton("选择文件")
        self.select_file_btn.setFixedWidth(100)
        layout.addWidget(self.select_file_btn)

        self.open_file_btn = QPushButton("打开文件")
        self.open_file_btn.setFixedWidth(100)
        layout.addWidget(self.open_file_btn)

        group.setLayout(layout)
        return group

    def create_param_group(self):
        """创建统一的参数设置区域"""
        group = QGroupBox("参数设置")
        layout = QGridLayout()
        layout.setSpacing(15)

        # 第一行：提取起始 和 搜索线程数
        layout.addWidget(QLabel("提取起始单元格:"), 0, 0)
        self.start_cell_edit = QLineEdit()  # 默认空
        self.start_cell_edit.setFixedWidth(80)
        layout.addWidget(self.start_cell_edit, 0, 1)

        layout.addWidget(QLabel("图片下载线程数:"), 0, 2)
        self.thread_spin = QSpinBox()
        self.thread_spin.setRange(1, 200)
        self.thread_spin.setValue(5)  # 默认5线程
        self.thread_spin.setFixedWidth(80)
        layout.addWidget(self.thread_spin, 0, 3)

        # 第二行：提取结束 和 启用网站筛选
        layout.addWidget(QLabel("提取结束单元格:"), 1, 0)
        self.end_cell_edit = QLineEdit()  # 默认空
        self.end_cell_edit.setFixedWidth(80)
        layout.addWidget(self.end_cell_edit, 1, 1)

        self.filter_check = QCheckBox("启用网站筛选")
        self.filter_check.setChecked(True)
        layout.addWidget(self.filter_check, 1, 2, 1, 2)

        # 第三行：插入起始
        layout.addWidget(QLabel("插入起始单元格:"), 2, 0)
        self.insert_cell_edit = QLineEdit()  # 默认空
        self.insert_cell_edit.setFixedWidth(80)
        layout.addWidget(self.insert_cell_edit, 2, 1)

        group.setLayout(layout)
        return group

    def create_content_log_group(self):
        """创建内容和日志区域 - 并排放置"""
        group = QGroupBox("内容与日志")
        layout = QHBoxLayout()
        layout.setSpacing(10)

        # 左列：待搜索内容
        content_widget = QWidget()
        content_layout = QVBoxLayout(content_widget)
        content_layout.setContentsMargins(0, 0, 0, 0)
        content_layout.addWidget(QLabel("待搜索内容:"))

        self.search_text = QTextEdit()
        self.search_text.setPlaceholderText("此处显示提取的待搜索内容")
        content_layout.addWidget(self.search_text)

        log_widget = QWidget()
        log_layout = QVBoxLayout(log_widget)
        log_layout.setContentsMargins(0, 0, 0, 0)
        log_layout.addWidget(QLabel("搜索日志:"))

        self.log_text = QTextEdit()
        self.log_text.setReadOnly(True)
        self.log_text.setStyleSheet("""
            QTextEdit {
                background-color: #f5f5f5;
                border: 1px solid #ccc;
                font-family: Consolas, 'Courier New', monospace;
                font-size: 9pt;
            }
        """)
        log_layout.addWidget(self.log_text)

        # 设置两个区域的宽度比例
        layout.addWidget(content_widget, 1)
        layout.addWidget(log_widget, 1)

        group.setLayout(layout)
        return group

    def create_button_group(self):
        """创建操作按钮区域"""
        group = QGroupBox("操作")
        layout = QHBoxLayout()
        layout.setSpacing(20)

        # 四个主要按钮
        self.extract_btn = QPushButton("提取件号")
        self.extract_btn.setFixedWidth(120)
        self.extract_btn.setFixedHeight(40)
        layout.addWidget(self.extract_btn)

        self.search_btn = QPushButton("搜索图片")
        self.search_btn.setFixedWidth(120)
        self.search_btn.setFixedHeight(40)
        layout.addWidget(self.search_btn)

        self.insert_btn = QPushButton("插入图片")
        self.insert_btn.setFixedWidth(120)
        self.insert_btn.setFixedHeight(40)
        layout.addWidget(self.insert_btn)

        self.auto_btn = QPushButton("一键操作")
        self.auto_btn.setFixedWidth(120)
        self.auto_btn.setFixedHeight(40)
        self.auto_btn.setStyleSheet("background-color: #4CAF50; color: white; font-weight: bold;")
        layout.addWidget(self.auto_btn)

        # 添加弹性空间使按钮居中
        layout.addStretch()

        group.setLayout(layout)
        return group

    def create_progress_group(self):
        """创建进度区域 - 完全重构"""
        group = QGroupBox("搜索进度")
        layout = QVBoxLayout()

        # 进度条
        self.progress_bar = QProgressBar()
        self.progress_bar.setValue(0)
        self.progress_bar.setTextVisible(True)
        self.progress_bar.setFormat("就绪")
        layout.addWidget(self.progress_bar)

        # 删除进度条底下的文字提示
        # 原来的self.progress_label已删除

        group.setLayout(layout)
        return group

    def setup_connections(self):
        """设置信号连接"""
        # 文件操作
        self.select_file_btn.clicked.connect(self.select_excel_file)
        self.open_file_btn.clicked.connect(self.open_excel_file)

        # 操作按钮
        self.extract_btn.clicked.connect(self.extract_content)
        self.search_btn.clicked.connect(self.start_search)
        self.insert_btn.clicked.connect(self.insert_images)
        self.auto_btn.clicked.connect(self.auto_process)

        # 搜索工作线程信号
        self.search_worker.search_started.connect(self.on_search_started)
        self.search_worker.item_completed.connect(self.on_item_completed)
        self.search_worker.search_finished.connect(self.on_search_finished)
        self.search_worker.search_error.connect(self.on_search_error)

    def select_excel_file(self):
        """选择Excel文件"""
        file_path, _ = QFileDialog.getOpenFileName(
            self, "选择Excel文件", "", "Excel文件 (*.xlsx)"
        )

        if file_path:
            self.file_path_edit.setText(file_path)

    def open_excel_file(self):
        """打开Excel文件"""
        file_path = self.file_path_edit.text()
        if file_path and os.path.exists(file_path):
            try:
                os.startfile(file_path)
            except Exception as e:
                QMessageBox.warning(self, "警告", f"无法打开文件: {str(e)}")
        else:
            QMessageBox.warning(self, "警告", "文件不存在")

    def extract_content(self):
        """提取Excel内容"""
        excel_file = self.file_path_edit.text()
        start_cell = self.start_cell_edit.text()
        end_cell = self.end_cell_edit.text()

        # 验证输入
        if not excel_file:
            QMessageBox.warning(self, "警告", "请选择Excel文件")
            return

        if not os.path.exists(excel_file):
            QMessageBox.warning(self, "警告", "文件不存在")
            return

        if not start_cell or not end_cell:
            QMessageBox.warning(self, "警告", "请输入提取范围")
            return

        try:
            # 提取数据
            data = extract_excel_data(start_cell, end_cell, excel_file)

            # 显示数据
            self.search_text.clear()
            if isinstance(data[0], list):
                for item in data:
                    self.search_text.append(" | ".join(item))
            else:
                for item in data:
                    self.search_text.append(item)

            # 检查空值
            if any("None" in str(item) for item in data):
                QMessageBox.warning(self, "警告", "提取出空值，请检查选择的文件以及输入的单元格是否正确！")

            self.add_log(f"提取完成，共{len(data)}个件号")

        except Exception as e:
            QMessageBox.critical(self, "错误", f"提取失败: {str(e)}")
            self.add_log(f"提取失败: {str(e)}", is_error=True)

    def start_search(self):
        """开始搜索图片"""
        # 获取搜索内容
        content = self.search_text.toPlainText()
        if not content.strip():
            QMessageBox.warning(self, "警告", "没有可搜索的内容，请先提取件号")
            return

        search_terms = content.split('\n')
        search_terms = [term.strip() for term in search_terms if term.strip()]

        if not search_terms:
            QMessageBox.warning(self, "警告", "没有可搜索的内容")
            return

        # 不清理日志，只添加开始标记（删除等号分割线）
        self.add_log(f"开始搜索，共 {len(search_terms)} 个件号，使用 {self.thread_spin.value()} 线程")

        # 重置进度
        self.completed_count = 0
        self.total_tasks = 0

        # 设置搜索参数
        self.search_worker.setup(
            search_terms=search_terms,
            max_workers=self.thread_spin.value(),
            enable_filter=self.filter_check.isChecked()
        )

        # 更新界面状态
        self.search_btn.setEnabled(False)
        self.extract_btn.setEnabled(False)

        # 开始搜索
        self.search_worker.start()

    def on_search_started(self, total_tasks):
        """搜索开始"""
        self.total_tasks = total_tasks
        self.completed_count = 0

        if total_tasks > 0:
            self.progress_bar.setMaximum(100)
            self.progress_bar.setValue(0)
            self.progress_bar.setFormat("0% (0/0)")
        else:
            self.progress_bar.setValue(0)
            self.progress_bar.setFormat("无任务")

    def on_item_completed(self, index, success, message):
        """单个任务完成"""
        self.completed_count += 1

        if self.total_tasks > 0:
            progress_percent = int((self.completed_count / self.total_tasks) * 100)
            self.progress_bar.setValue(progress_percent)
            self.progress_bar.setFormat(f"{progress_percent}% ({self.completed_count}/{self.total_tasks})")

        # 添加日志
        terms = self.search_text.toPlainText().split('\n')
        if index < len(terms):
            term = terms[index]
            timestamp = time.strftime("%H:%M:%S")

            if success:
                log_msg = f"[{timestamp}] ✓ {term}"
                color = "green"
            else:
                log_msg = f"[{timestamp}] ✗ {term} - {message}"
                color = "red"

            cursor = self.log_text.textCursor()
            cursor.movePosition(QTextCursor.MoveOperation.End)

            if success:
                html = f'<font color="{color}">{log_msg}</font>'
            else:
                html = f'<font color="{color}"><b>{log_msg}</b></font>'

            cursor.insertHtml(html + "<br>")
            self.log_text.ensureCursorVisible()

    def on_search_finished(self, failed_items):
        """搜索完成"""
        self.search_btn.setEnabled(True)
        self.extract_btn.setEnabled(True)

        # 完成时进度条设为100%
        self.progress_bar.setValue(100)
        self.progress_bar.setFormat("100% (完成)")

        if failed_items:
            # 使用黄色显示失败消息
            self.add_log(f"搜索完成，{len(failed_items)} 个件号失败", is_warning=True)
            QMessageBox.warning(self, "完成",
                                f"搜索完成！但有 {len(failed_items)} 个件号图片无法搜到")
        else:
            self.add_log("搜索完成，所有图片已成功下载")
            QMessageBox.information(self, "完成", "搜索完成！")

    def on_search_error(self, error_message):
        """搜索出错"""
        self.search_btn.setEnabled(True)
        self.extract_btn.setEnabled(True)

        self.progress_bar.setValue(0)
        self.progress_bar.setFormat("错误")

        self.add_log(f"搜索失败: {error_message}", is_error=True)
        QMessageBox.critical(self, "错误", error_message)

    def add_log(self, message, is_error=False, is_warning=False):
        """添加普通日志"""
        timestamp = time.strftime("%H:%M:%S")

        if is_error:
            log_msg = f"[{timestamp}] ✗ {message}"
            color = "red"
        elif is_warning:
            log_msg = f"[{timestamp}] ⚠ {message}"
            color = "#FFA500"  # 橙色
        else:
            log_msg = f"[{timestamp}] ● {message}"
            color = "blue"

        cursor = self.log_text.textCursor()
        cursor.movePosition(QTextCursor.MoveOperation.End)

        if is_error or is_warning:
            html = f'<font color="{color}"><b>{log_msg}</b></font>'
        else:
            html = f'<font color="{color}">{log_msg}</font>'

        cursor.insertHtml(html + "<br>")
        self.log_text.ensureCursorVisible()

    def insert_images(self):
        """插入图片到Excel"""
        excel_file = self.file_path_edit.text()
        insert_cell = self.insert_cell_edit.text()

        # 验证输入
        if not excel_file:
            QMessageBox.warning(self, "警告", "请选择Excel文件")
            return

        if not os.path.exists(excel_file):
            QMessageBox.warning(self, "警告", "文件不存在")
            return

        if not insert_cell:
            QMessageBox.warning(self, "警告", "请输入插入起始单元格")
            return

        # 检查图片目录是否存在
        if not os.path.exists(IMG_DIR):
            QMessageBox.warning(self, "警告", "未找到图片文件夹，请先搜索图片")
            return

        # 计算图片数量
        start_cell = self.start_cell_edit.text()
        end_cell = self.end_cell_edit.text()

        if not start_cell or not end_cell:
            QMessageBox.warning(self, "警告", "请输入提取范围以计算图片数量")
            return

        try:
            # 提取单元格行号
            start_row = int(''.join(filter(str.isdigit, start_cell)))
            end_row = int(''.join(filter(str.isdigit, end_cell)))
            image_count = end_row - start_row + 1

            self.add_log(f"开始插入图片，从 {insert_cell} 开始，共 {image_count} 张")

            # 插入图片
            error_count = insert_images_to_excel(insert_cell, image_count, excel_file)

            if error_count > 0:
                self.add_log(f"插入完成，{error_count} 个图片插入失败", is_warning=True)
                QMessageBox.warning(self, "完成", f"有 {error_count} 个图片插入失败！其余插入成功。")
            else:
                self.add_log("插入完成，所有图片已成功插入")
                QMessageBox.information(self, "完成", "插入完成！")

        except Exception as e:
            QMessageBox.critical(self, "错误", f"插入失败: {str(e)}")
            self.add_log(f"插入失败: {str(e)}", is_error=True)

    def auto_process(self):
        """一键操作：提取->搜索->插入"""
        # 第一步：提取内容
        self.extract_content()

        # 等待提取完成
        QApplication.processEvents()
        QTimer.singleShot(500, self.auto_process_step2)

    def auto_process_step2(self):
        """一键操作第二步：搜索"""
        content = self.search_text.toPlainText()
        if not content.strip():
            return

        self.start_search()

        # 当搜索完成时自动插入
        try:
            self.search_worker.search_finished.disconnect(self.auto_insert_after_search)
        except:
            pass

        self.search_worker.search_finished.connect(self.auto_insert_after_search)

    def auto_insert_after_search(self, failed_items):
        """搜索完成后自动插入"""
        try:
            self.search_worker.search_finished.disconnect(self.auto_insert_after_search)
        except:
            pass

        QTimer.singleShot(1000, self.insert_images)

    def closeEvent(self, event):
        """关闭窗口事件"""
        if self.search_worker.isRunning():
            self.search_worker.terminate()
            self.search_worker.wait()
        event.accept()


def main():
    """主函数"""
    app = QApplication(sys.argv)
    app.setStyle('Fusion')
    window = ExcelToolsGUI()
    window.show()
    sys.exit(app.exec())


if __name__ == '__main__':
    main()